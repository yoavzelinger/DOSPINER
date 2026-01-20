print("Starting import")
from argparse import ArgumentParser
import os
from openpyxl import load_workbook

import pandas as pd

import Tester.TesterConstants as tester_constants

parser = ArgumentParser(description="Run all tests")
parser.add_argument("-i", "--input", type=str, help=f"Input folder prefix (after the temp), default is None ({tester_constants.TEMP_OUTPUT_DIRECTORY_NAME})", default="")
parser.add_argument("-o", "--output", type=str, help=f"Output file name prefix, default is the result_{tester_constants.DEFAULT_RESULTS_FILENAME_PREFIX}_<TIMESTAMP>", default=tester_constants.DEFAULT_RESULTS_FILENAME_EXTENDED_PREFIX)
parser.add_argument("-r", "--raw", action="store_true", help="Whether to create a file with the raw results. Note that the raw results file can be very big and infeasible to store.")
args = parser.parse_args()

if args.output == tester_constants.DEFAULT_RESULTS_FILENAME_EXTENDED_PREFIX and args.input:
    args.output = args.input

aggregating_functions_dict = {tester_constants.AGGREGATED_TESTS_COUNT_COLUMN: "count"}
aggregating_functions_dict |= {metric_column_name: "sum" for metric_column_name in tester_constants.AGGREGATED_METRICS_COLUMNS}

output_df = pd.DataFrame(columns=tester_constants.GROUP_BY_COLUMN_NAMES + tester_constants.EXTENDED_METRICS_COLUMN_NAMES).astype(tester_constants.GROUP_BY_COLUMNS | tester_constants.EXTENDED_METRICS_COLUMNS).set_index(tester_constants.GROUP_BY_COLUMN_NAMES)
raw_df = pd.DataFrame(columns=tester_constants.RAW_RESULTS_COLUMN_NAMES).astype(tester_constants.RAW_RESULTS_COLUMNS)

temp_output_directory_full_path = tester_constants.TEMP_OUTPUT_DIRECTORY_FULL_PATH
if args.input:
   temp_output_directory_full_path = f"{temp_output_directory_full_path}_{args.input}"
assert os.path.isdir(temp_output_directory_full_path), f"{temp_output_directory_full_path} does not exists"
print(f"Merging all temp files from {temp_output_directory_full_path}")

for current_file_index, current_file_name in enumerate(os.listdir(temp_output_directory_full_path), 1):
    print("Working on file", current_file_index, ":", current_file_name)
    if not current_file_name.startswith(tester_constants.RESULTS_FILE_NAME_PREFIX):
        continue
    
    current_results_df = None
    with open(os.path.join(temp_output_directory_full_path, current_file_name), "r") as current_file:
        current_results_df = pd.read_csv(current_file, dtype=tester_constants.RAW_RESULTS_COLUMNS)
    if args.raw:
        raw_df = pd.concat([raw_df, current_results_df], ignore_index=True)        
    current_group_by_df = current_results_df.groupby(tester_constants.GROUP_BY_COLUMN_NAMES).agg(aggregating_functions_dict)
    # check that no column contains empty values
    if current_group_by_df.isnull().values.any():
        raise ValueError(f"Empty values found in {current_file_name}. Please check the input data.")
    current_group_by_df.rename(columns={tester_constants.AGGREGATED_TESTS_COUNT_COLUMN: tester_constants.TESTS_COUNTS_COLUMN_NAME}, inplace=True)
    if len(current_group_by_df.columns) == len(tester_constants.EXTENDED_METRICS_COLUMN_NAMES) and not (all(current_group_by_df.columns == output_df.columns)):
        raise AssertionError(
            f"Columns mismatch in {current_file_name}:\n"
            f"Mismatched columns by position: {[(i, c1, c2) for i, (c1, c2) in enumerate(zip(current_group_by_df.columns, output_df.columns)) if c1 != c2]}\n"
            f"Extra columns in current_group_by_df: {[c for c in current_group_by_df.columns if c not in output_df.columns]}\n"
            f"Extra columns in output_df: {[c for c in output_df.columns if c not in current_group_by_df.columns]}"
        )
    if not (all(current_group_by_df.dtypes == output_df.dtypes)):
        current_dict_dtypes = dict(current_group_by_df.dtypes)
        raise AssertionError(
            f"Dtypes mismatch in {current_file_name}:\n"
            f"Extra dtypes in output_df: {[f'{output_column} (output[{output_dtype}] != current[{current_dict_dtypes[output_column]}])' for output_column, output_dtype in dict(output_df.dtypes).items() if output_dtype != current_dict_dtypes[output_column]]}"
        )
    if current_group_by_df.index.names != output_df.index.names:
        raise AssertionError(
            f"Index names mismatch in {current_file_name}:\n"
            f"Current index names: {current_group_by_df.index.names}, Output index names: {output_df.index.names}"
        )

    output_df = output_df.add(current_group_by_df, fill_value=0).astype(tester_constants.EXTENDED_METRICS_COLUMNS)



output_df = output_df[tester_constants.EXTENDED_METRICS_COLUMN_NAMES]

output_full_path_prefix = os.path.join(tester_constants.OUTPUT_DIRECTORY_FULL_PATH, f"{tester_constants.RESULTS_FILE_NAME_PREFIX}_{args.output}")

merged_output_full_path = f"{output_full_path_prefix}.xlsx"
excel_writer_arguments = {
    "path": merged_output_full_path,
    "mode": "w",
    "engine": "openpyxl"
}
if os.path.exists(merged_output_full_path):
    excel_writer_arguments["mode"] = "a"
    excel_writer_arguments["if_sheet_exists"] = "replace"
with pd.ExcelWriter(**excel_writer_arguments) as excel_writer:
    if args.raw:
        raw_df.to_excel(excel_writer, sheet_name=tester_constants.RAW_RESULTS_SHEET_NAME, merge_cells=False, index=False)
    output_df.to_excel(excel_writer, sheet_name=tester_constants.MERGED_RESULTS_SHEET_NAME, merge_cells=False)
output_workbook = load_workbook(merged_output_full_path)
if args.raw:
    output_workbook[tester_constants.RAW_RESULTS_SHEET_NAME].sheet_view.rightToLeft = True
output_workbook[tester_constants.MERGED_RESULTS_SHEET_NAME].sheet_view.rightToLeft = True
output_workbook.save(merged_output_full_path)

print(f"Results saved to {output_full_path_prefix}")